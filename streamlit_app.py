"""GR BMMS — live imbalance monitor: proxy price, system state, mFRR prices.

One page, four stacked live charts on the quarter cadence — actual data first,
the approximation last:

  1) mFRR PRICES (€/MWh) — activated balancing-energy prices Up / Down. Actual
     data. Y-view fixed to −100…250 €/MWh so sparse spikes don't zoom the chart
     out — they exit the top; drag-zoom to inspect, double-click to reset.
  2) mFRR NET ACTIVATED ENERGY (MWh per 15-min quarter, up − down, mFRR ONLY) —
     actual data, red = net up, green = net down.
  3) mFRR + aFRR NET ACTIVATED ENERGY (MWh per 15-min quarter, up − down) —
     actual data: the full ENTSO-E activation net that feeds the estimate.
  4) ESTIMATED SYSTEM STATE (MWh per 15-min quarter, short/long) — approximation,
     standalone. Positive = short (up), negative = long (down).

Efficiency: one canonical cached fetch per source shared by every viewer —
prices + bids every 2 min (3 ENTSO-E requests), ISP publications every 15 min
(workbooks downloaded only when newly published).

Token resolution:  st.secrets["ENTSOE_TOKEN"] -> env ENTSOE_TOKEN -> "entsoe Token.txt"
Correction artifacts (sysdev_*.csv) come from sysdev_estimator.ipynb — refresh weekly.
"""

import io
import json
import os
import re
import urllib.request

import numpy as np
import openpyxl
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from entsoe import EntsoePandasClient
from entsoe.exceptions import NoMatchingDataError

from entsoe_bids_fix import query_aggregated_bids_fixed

ZONE = "GR"
QUARTER = pd.Timedelta(minutes=15)
CACHE_TTL_S = 120              # ENTSO-E fetches: once per 2 min, shared by all viewers
SURPLUS_TTL_S = 900            # ISP publications checked every 15 min
REFRESH_S = 60                 # redraw cadence
FETCH_WINDOW_H = 49            # canonical window shared by every viewer / window setting
QPH = 4.0                      # quarters per hour — raw ENTSO-E/ISP MW ÷ QPH = MWh per 15-min quarter
PRICE_VIEW = (-100, 250)       # fixed default €/MWh view — spikes exit the top, zoom to inspect

APP_DIR = os.path.dirname(os.path.abspath(__file__))
FONT = 'system-ui, -apple-system, "Segoe UI", sans-serif'
PALETTE = dict(
    up="#e66767", down="#008300",
    short_fill="rgba(230,103,103,0.30)", long_fill="rgba(0,131,0,0.30)",
    line="#ffffff",
    ink="#ffffff", ink2="#c3c2b7", muted="#898781",
    grid="#2c2c2a", baseline="#383835",
    pending="rgba(137,135,129,0.18)",
    now_band="rgba(57,135,229,0.16)",
    last_band="rgba(201,133,0,0.22)",
    separator="rgba(137,135,129,0.60)",
    hover_bg="#1a1a19", ring="rgba(255,255,255,0.10)",
)

ADMIE_API = "https://www.admie.gr/getOperationMarketFilewRange?dateStart={s}&dateEnd={e}&FileCategory={c}"
ISP_CATS = ["ISP1ISPResults", "ISP2ISPResults", "ISP3ISPResults", "AdhocISPResults"]


def _token() -> str:
    tok = ""
    try:
        tok = st.secrets.get("ENTSOE_TOKEN", "")
    except Exception:
        pass
    tok = tok or os.environ.get("ENTSOE_TOKEN", "")
    if not tok:
        path = os.path.join(APP_DIR, "entsoe Token.txt")
        if os.path.exists(path):
            tok = open(path).read().strip()
    if not tok or "PASTE_YOUR" in tok:
        raise RuntimeError(
            "No ENTSO-E API token found. Set ENTSOE_TOKEN in Streamlit secrets / "
            "environment, or paste it into 'entsoe Token.txt' next to this app."
        )
    return tok


@st.cache_resource
def load_artifacts():
    coefs = pd.read_csv(os.path.join(APP_DIR, "sysdev_surplus_coefs.csv"), index_col=0)["value"]
    left = pd.read_csv(os.path.join(APP_DIR, "sysdev_leftover_profile.csv"), index_col=0).iloc[:, 0]
    v1 = pd.read_csv(os.path.join(APP_DIR, "sysdev_correction_profile.csv"), index_col=0).iloc[:, 0]
    return float(coefs["alpha"]), float(coefs["beta"]), left, v1


# ---------------- data: prices (A97), bids net (A47+A67), ISP surplus ----------------
@st.cache_data(ttl=CACHE_TTL_S, show_spinner=False)
def fetch_prices() -> pd.DataFrame:
    """GR mFRR Up/Down activated-energy prices, CET-naive quarter index."""
    client = EntsoePandasClient(api_key=_token())
    now = pd.Timestamp.now(tz="CET")
    start = (now - pd.Timedelta(hours=FETCH_WINDOW_H)).floor("h")
    end = now.ceil("h")
    try:
        r = client.query_activated_balancing_energy_prices(
            ZONE, start=start, end=end, business_type="A97")
    except NoMatchingDataError:
        return pd.DataFrame()
    r.index.name = "ts"
    r = r.reset_index()
    p = r.pivot_table(index="ts", columns="Direction", values="Price", aggfunc="first")
    p.index = p.index.tz_convert("CET").tz_localize(None)
    return p.sort_index()


@st.cache_data(ttl=CACHE_TTL_S, show_spinner=False)
def fetch_net() -> pd.DataFrame:
    """ENTSO-E net activated energy, MWh per quarter, CET-naive, trimmed to the
    activation frontier. Columns: mfrr_net (mFRR only — actual data for chart 2)
    and total_net (mFRR+aFRR — feeds the state estimate)."""
    client = EntsoePandasClient(api_key=_token())
    now = pd.Timestamp.now(tz="CET")
    start = (now - pd.Timedelta(hours=FETCH_WINDOW_H)).floor("h")
    end = now.ceil("h")
    nets = {}
    for pt, key in (("A47", "mfrr_net"), ("A67", "afrr_net")):
        try:
            raw = query_aggregated_bids_fixed(client, ZONE, process_type=pt, start=start, end=end)
        except NoMatchingDataError:
            continue
        act = raw.xs("Activated", axis=1, level="unit").apply(pd.to_numeric, errors="coerce") / QPH
        dirs = act.columns.get_level_values("direction")
        up = act.loc[:, dirs == "Up"].sum(axis=1, min_count=1) if (dirs == "Up").any() else pd.Series(np.nan, index=act.index)
        dn = act.loc[:, dirs == "Down"].sum(axis=1, min_count=1) if (dirs == "Down").any() else pd.Series(np.nan, index=act.index)
        both = up.isna() & dn.isna()
        net = up.fillna(0) - dn.fillna(0)
        net[both] = np.nan
        net.index = net.index.tz_convert("CET").tz_localize(None)
        nets[key] = net.sort_index()
    if not nets:
        return pd.DataFrame()
    df = pd.DataFrame(nets)
    parts = [df[c] for c in ("mfrr_net", "afrr_net") if c in df.columns]
    total = parts[0].fillna(0) if parts else pd.Series(dtype=float)
    for p in parts[1:]:
        total = total.add(p.fillna(0), fill_value=0)
    total[df.isna().all(axis=1)] = np.nan
    df["total_net"] = total
    last = df["total_net"].last_valid_index()
    return df.loc[:last] if last is not None else pd.DataFrame()


def _parse_surplus_xlsx(data: bytes) -> pd.Series | None:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheets = [sh for sh in wb.sheetnames if sh.endswith("_ISP")]
        if not sheets:
            wb.close()
            return None
        rows = [r for r in wb[sheets[0]].iter_rows(values_only=True)]
        wb.close()
        ts = None
        for r in rows:
            # timestamps live in cols 1..96 (00:00..23:45); col 0 is the row label,
            # col 97 is the daily TOTAL. Start at r[1:] so the 00:00 quarter is kept.
            cells = [c for c in r[1:] if c is not None and hasattr(c, "hour")]
            if len(cells) > 50:
                ts = [pd.Timestamp(c) for c in r[1:] if c is not None and hasattr(c, "hour")]
                break
        if not ts:
            return None
        for r in rows:
            if r[0] == "Energy Surplus":
                vals = pd.to_numeric(pd.Series(list(r[1:1 + len(ts)])), errors="coerce")
                return pd.Series(vals.values, index=pd.DatetimeIndex(ts))
    except Exception:
        return None
    return None


@st.cache_data(ttl=86400, max_entries=48, show_spinner=False)
def _isp_series(url: str) -> pd.Series | None:
    """One ISP workbook, cached by URL for a day — downloads happen only for
    genuinely new publications. A download error propagates (st.cache_data does NOT
    cache exceptions) so a transient failure is retried next cycle instead of being
    pinned as 'no data' for 24 h; a workbook that has no Energy Surplus row returns
    None (safe to cache — that URL's content is immutable)."""
    data = urllib.request.urlopen(
        urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"}), timeout=90).read()
    return _parse_surplus_xlsx(data)


@st.cache_data(ttl=SURPLUS_TTL_S, show_spinner=False)
def fetch_surplus() -> pd.Series:
    """Scheduled ISP Energy Surplus (yesterday+today), MWh per quarter (negative),
    latest-run-wins across ISP1/2/3/Ad-hoc."""
    now = pd.Timestamp.now(tz="CET").tz_localize(None)
    d0, d1 = now.normalize() - pd.Timedelta(days=1), now.normalize()
    pat = re.compile(r"/([^/]+\.xlsx)$", re.I)
    stage_of = {"ISP1": 1, "ISP2": 2, "ISP3": 3, "Adho": 4}
    entries = []
    for cat in ISP_CATS:
        try:
            req = urllib.request.Request(
                ADMIE_API.format(s=d0.strftime("%Y-%m-%d"), e=d1.strftime("%Y-%m-%d"), c=cat),
                headers={"User-Agent": "Mozilla/5.0"})
            listing = json.loads(urllib.request.urlopen(req, timeout=45).read().decode("utf-8"))
        except Exception:
            continue
        best = {}
        for it in listing:
            m = pat.search(it.get("file_path", ""))
            if not m:
                continue
            fn = m.group(1)
            v = re.search(r"_(\d+)\.xlsx$", fn)
            key = fn[:8]
            ver = int(v.group(1)) if v else 0
            if key not in best or ver > best[key][0]:
                best[key] = (ver, it["file_path"])
        for day, (ver, url) in best.items():
            entries.append((day, stage_of.get(cat[:4], 9), ver, url))
    out = None
    for day, stage, ver, url in sorted(entries):
        try:
            s = _isp_series(url)
        except Exception:
            continue                              # transient download error — not cached, retried next cycle
        if s is None:
            continue
        out = s if out is None else out
        out = s.combine_first(out)
        out.update(s.dropna())
    if out is None:
        return pd.Series(dtype=float)
    return (out / QPH).sort_index()


def build_state(net_mwh: pd.Series, surplus_mwh: pd.Series) -> pd.DataFrame:
    """state = (net + α + β(−surplus) + leftover[hour], v1[hour] fallback) + surplus.
    CET-naive index (profiles are CET-hour keyed). Columns: net, surplus, state (MWh)."""
    alpha, beta, left, v1 = load_artifacts()
    f = pd.DataFrame({"net": net_mwh})
    f["surplus"] = surplus_mwh.reindex(f.index) if len(surplus_mwh) else np.nan
    hours = pd.Series(f.index.hour, index=f.index)
    dev = (f["net"] + alpha + beta * (-f["surplus"]) + hours.map(left)).fillna(f["net"] + hours.map(v1))
    f["state"] = dev + f["surplus"].fillna(0)
    return f


# ---------------- charts ----------------
def _zones(fig, s_index_max, now, c):
    q_now = now.floor("15min")
    if pd.notna(s_index_max):                     # skip data-frontier bands when the series is empty (NaT)
        last_end = s_index_max + QUARTER
        pending_n = max(0, int((q_now - last_end) / QUARTER))
        fig.add_vrect(x0=s_index_max, x1=last_end, fillcolor=c["last_band"], line_width=0, layer="below")
        if pending_n > 0:
            fig.add_vrect(x0=last_end, x1=q_now, fillcolor=c["pending"], line_width=0, layer="below")
    fig.add_vrect(x0=q_now, x1=q_now + QUARTER, fillcolor=c["now_band"], line_width=0, layer="below")
    fig.add_vline(x=now, line=dict(color=c["ink"], width=1.5, dash="dash"))


def _layout(fig, now, window_h, ylab, c, height=300):
    pad = pd.Timedelta(minutes=max(30, int(window_h * 60 * 0.07)))
    dtick_h = {6: 1, 12: 2, 24: 4, 48: 8}.get(window_h, 4)
    fig.update_layout(
        template="none", paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        font=dict(family=FONT, size=12, color=c["ink2"]),
        margin=dict(l=4, r=8, t=26, b=8), height=height, hovermode="x", hoverdistance=-1,
        hoverlabel=dict(bgcolor=c["hover_bg"], bordercolor=c["ring"],
                        font=dict(family=FONT, size=12, color=c["ink"])),
        legend=dict(orientation="h", x=0, y=1.18, yanchor="bottom", font=dict(size=12, color=c["ink2"])),
        xaxis=dict(range=[now - pd.Timedelta(hours=window_h), now + pad],
                   showgrid=True, gridcolor=c["grid"], gridwidth=1,
                   dtick=dtick_h * 3600 * 1000, tickformat="%H:%M", automargin=True,
                   tickfont=dict(color=c["muted"]), hoverformat="%d %b %H:%M",
                   linecolor=c["baseline"], showline=True),
        yaxis=dict(title=dict(text=ylab, font=dict(color=c["muted"], size=12)),
                   showgrid=True, gridcolor=c["grid"], gridwidth=1, automargin=True,
                   zeroline=True, zerolinecolor=c["baseline"], zerolinewidth=1,
                   tickfont=dict(color=c["muted"])),
    )
    return fig


def _step_xy(series):
    s = series.dropna()
    return list(s.index) + [s.index[-1] + QUARTER], list(s.values) + [s.values[-1]]


def _fmt_q(ts) -> str:
    return f"{ts:%H:%M}–{(ts + QUARTER):%H:%M}"


def _fmt_mwh(v) -> str:
    return "—" if pd.isna(v) else f"{v:+,.1f} MWh"


def _fmt_px(v) -> str:
    return "—" if pd.isna(v) else f"{v:,.2f} €/MWh"


def _hover_quarters(now, window_h):
    lo = (now - pd.Timedelta(hours=window_h)).floor("15min")
    return pd.date_range(lo, now.floor("15min"), freq="15min")


def _hover_trace(quarters, ypos, rows, template):
    """The only hover-active trace on each chart: invisible markers at quarter
    MIDPOINTS, so the cursor always resolves to the quarter it is inside — line
    traces skip hover, otherwise plotly snaps to the nearest point and pending
    quarters echo the previous quarter's value. Rows carry the 13:00–13:15
    label and pre-formatted values ("—" until a quarter is published)."""
    return go.Scatter(
        x=quarters + pd.Timedelta(minutes=7.5), y=ypos, mode="markers",
        marker=dict(size=1, color="rgba(0,0,0,0)"), showlegend=False,
        customdata=rows, hovertemplate=template, hoverlabel=dict(align="left"),
    )


def chart_signed(series_mwh: pd.Series, name: str, ylab: str, now, window_h, c) -> go.Figure:
    """Sign-filled step chart (red above zero / green below) for MWh-per-quarter series."""
    fig = go.Figure()
    _zones(fig, series_mwh.dropna().index.max(), now, c)
    x, y = _step_xy(series_mwh)
    ys = pd.Series(y, index=x)
    fig.add_trace(go.Scatter(x=x, y=ys.clip(lower=0), mode="lines", line=dict(width=0, shape="hv"),
                             fill="tozeroy", fillcolor=c["short_fill"], hoverinfo="skip", showlegend=False))
    fig.add_trace(go.Scatter(x=x, y=ys.clip(upper=0), mode="lines", line=dict(width=0, shape="hv"),
                             fill="tozeroy", fillcolor=c["long_fill"], hoverinfo="skip", showlegend=False))
    fig.add_trace(go.Scatter(x=x, y=y, mode="lines", name=name,
                             line=dict(color=c["line"], width=2, shape="hv"),
                             hoverinfo="skip"))
    quarters = _hover_quarters(now, window_h)
    main_q = series_mwh.reindex(quarters)
    rows = [[_fmt_q(q), _fmt_mwh(v)] for q, v in zip(quarters, main_q)]
    template = (f"<b>%{{customdata[0]}}</b><br>"
                f"<span style='color:{c['line']}'>●</span> {name}  %{{customdata[1]}}<extra></extra>")
    fig.add_trace(_hover_trace(quarters, main_q.fillna(0.0), rows, template))
    last = float(series_mwh.dropna().iloc[-1])
    fig.add_annotation(xref="x domain", x=0.995, y=last, showarrow=False,
                       text=f"{last:+,.1f} MWh", xanchor="right", yshift=10,
                       font=dict(size=12, color=c["ink"], family=FONT))
    return _layout(fig, now, window_h, ylab, c)


def chart_prices(up: pd.Series, dn: pd.Series, now, window_h, c) -> go.Figure:
    fig = go.Figure()
    both = pd.concat([up, dn]).dropna()
    _zones(fig, both.index.max(), now, c)
    labels = []
    for s, color, name in ((up, c["up"], "Up"), (dn, c["down"], "Down")):
        sd = s.dropna()
        if sd.empty:
            continue
        x, y = _step_xy(sd)
        fig.add_trace(go.Scatter(x=x, y=y, mode="lines", name=f"mFRR {name}",
                                 line=dict(color=color, width=2, shape="hv"),
                                 hoverinfo="skip"))
        labels.append((name, float(sd.iloc[-1])))
    shift = {0: 0, 1: 0}
    if len(labels) == 2:
        hi = 0 if labels[0][1] >= labels[1][1] else 1
        shift = {hi: 10, 1 - hi: -10}
    for i, (name, val) in enumerate(labels):
        # keep the label inside the clamped view even when the price spikes out of it
        fig.add_annotation(xref="x domain", x=0.995, y=min(max(val, PRICE_VIEW[0] + 10), PRICE_VIEW[1] - 10),
                           text=f"{name} {val:,.1f}", showarrow=False, xanchor="right", yshift=shift[i],
                           font=dict(size=12, color=c["ink2"], family=FONT))
    quarters = _hover_quarters(now, window_h)
    upq, dnq = up.reindex(quarters), dn.reindex(quarters)
    rows = [[_fmt_q(q), _fmt_px(u), _fmt_px(d)] for q, u, d in zip(quarters, upq, dnq)]
    ypos = upq.combine_first(dnq).fillna(0.0).clip(PRICE_VIEW[0] + 10, PRICE_VIEW[1] - 10)
    fig.add_trace(_hover_trace(
        quarters, ypos, rows,
        f"<b>%{{customdata[0]}}</b><br>"
        f"<span style='color:{c['up']}'>●</span> Up  %{{customdata[1]}}<br>"
        f"<span style='color:{c['down']}'>●</span> Down  %{{customdata[2]}}<extra></extra>"))
    fig = _layout(fig, now, window_h, "€/MWh", c)
    # fixed default view: spikes run off the top instead of zooming the chart out;
    # drag to zoom/pan for the full picture, double-click restores this band
    fig.update_yaxes(range=list(PRICE_VIEW), autorange=False)
    return fig


def _state_word(v: float) -> str:
    return "SHORT" if v > 0 else "LONG"


# ---------------- live view ----------------
@st.fragment(run_every=REFRESH_S)
def live_view(window_h: int, tz: str, tz_short: str):
    try:
        load_artifacts()
    except Exception:
        st.error("Correction artifacts missing — copy the three sysdev_*.csv files next to this app.")
        return
    try:
        prices = fetch_prices()
        net = fetch_net()
    except RuntimeError as e:
        st.error(str(e))
        return
    except Exception as e:
        detail = type(e).__name__
        resp = getattr(e, "response", None)
        if resp is not None and getattr(resp, "status_code", None):
            detail = f"HTTP {resp.status_code}"
            if resp.status_code == 401:
                detail += " — ENTSO-E rejected the API token (check Secrets)."
            elif resp.status_code == 429:
                detail += " — rate limited; will keep retrying."
        st.warning(f"ENTSO-E request failed ({detail}). Retrying in {REFRESH_S}s.")
        return
    if prices.empty or net.empty:
        st.warning("ENTSO-E returned no data yet. Retrying automatically.")
        return
    try:
        surplus = fetch_surplus()
    except Exception:
        surplus = pd.Series(dtype=float)

    est = build_state(net["total_net"], surplus)         # CET-naive, MWh

    # display conversion
    shift = pd.Timedelta(hours=1) if tz == "Europe/Athens" else pd.Timedelta(0)
    now = pd.Timestamp.now(tz=tz).tz_localize(None)
    lo = now - pd.Timedelta(hours=window_h)

    def dsp(s):
        s = s.copy()
        s.index = s.index + shift
        return s[s.index >= lo]

    up_d, dn_d = dsp(prices.get("Up", pd.Series(dtype=float))), dsp(prices.get("Down", pd.Series(dtype=float)))
    mfrr_mwh = dsp(net.get("mfrr_net", pd.Series(dtype=float)))
    state_mwh, total_mwh = dsp(est["state"]), dsp(est["net"])
    sur_mwh = dsp(est["surplus"])
    if mfrr_mwh.dropna().empty or total_mwh.dropna().empty or state_mwh.dropna().empty:
        st.warning("No overlapping data inside the selected window yet.")
        return

    q_now = now.floor("15min")
    e_last = mfrr_mwh.dropna().index.max()
    n_now = float(mfrr_mwh.dropna().iloc[-1])
    s_now = float(state_mwh.dropna().iloc[-1])
    s_prev = float(state_mwh.dropna().iloc[-2]) if state_mwh.notna().sum() > 1 else None
    lag_q = max(0, int((q_now - (e_last + QUARTER)) / QUARTER))

    def px_sub(s):
        sd = s.dropna()
        return f"quarter {_fmt_q(sd.index.max())}" if len(sd) else "no price data yet"

    c = PALETTE
    tiles = [
        ("mFRR Up", "—" if up_d.dropna().empty else f"{float(up_d.dropna().iloc[-1]):,.2f} €/MWh",
         px_sub(up_d)),
        ("mFRR Down", "—" if dn_d.dropna().empty else f"{float(dn_d.dropna().iloc[-1]):,.2f} €/MWh",
         px_sub(dn_d)),
        ("mFRR net energy", f"{n_now:+,.1f} MWh",
         f"quarter {_fmt_q(e_last)}"),
        ("System state (est.)", f"{s_now:+,.1f} MWh · {_state_word(s_now)}",
         "" if s_prev is None else f"Δ {s_now - s_prev:+,.1f} MWh vs prev quarter"),
        ("Energy data lag", "up to date" if lag_q == 0 else f"{lag_q} × 15 min",
         f"now in {_fmt_q(q_now)} {tz_short}"),
    ]
    tile_html = "".join(
        f'<div style="flex:1 1 150px;min-width:150px;padding:10px 14px;'
        f'border:1px solid {c["ring"]};border-radius:10px;">'
        f'<div style="font-size:11px;letter-spacing:.04em;text-transform:uppercase;'
        f'color:{c["muted"]};margin-bottom:2px;">{label}</div>'
        f'<div style="font-size:22px;font-weight:600;color:{c["ink"]};'
        f'font-variant-numeric:tabular-nums;line-height:1.25;">{value}</div>'
        f'<div style="font-size:12px;color:{c["ink2"]};margin-top:2px;">{sub}&nbsp;</div></div>'
        for label, value, sub in tiles
    )
    st.markdown(f'<div style="display:flex;flex-wrap:wrap;gap:12px;margin-bottom:4px;'
                f'font-family:{FONT};">{tile_html}</div>', unsafe_allow_html=True)

    st.markdown("##### 1 · mFRR activated balancing-energy prices")
    st.plotly_chart(chart_prices(up_d, dn_d, now, window_h, c),
                    width="stretch", config={"displayModeBar": False})
    st.markdown("##### 2 · mFRR net activated energy (up − down)")
    st.plotly_chart(chart_signed(mfrr_mwh, "mFRR net", "MWh per quarter (+up / −down)",
                                 now, window_h, c),
                    width="stretch", config={"displayModeBar": False})
    st.markdown("##### 3 · mFRR + aFRR net activated energy (up − down)")
    st.plotly_chart(chart_signed(total_mwh, "mFRR + aFRR net", "MWh per quarter (+up / −down)",
                                 now, window_h, c),
                    width="stretch", config={"displayModeBar": False})
    st.markdown("##### 4 · Estimated system state (short/long)")
    st.plotly_chart(chart_signed(state_mwh, "system state (est.)",
                                 "MWh per quarter (+short / −long)", now, window_h, c),
                    width="stretch", config={"displayModeBar": False})

    with st.expander("Data table & CSV download"):
        t = pd.DataFrame({
            "mFRR Up (€/MWh)": up_d.round(2),
            "mFRR Down (€/MWh)": dn_d.round(2),
            "mFRR net (MWh)": mfrr_mwh.round(2),
            "mFRR+aFRR net (MWh)": total_mwh.round(2),
            "State (MWh)": state_mwh.round(2),
            "Surplus (MWh)": sur_mwh.round(2),
        }).sort_index(ascending=False)
        t.insert(0, "Quarter", [f"{ts:%H:%M}–{(ts + QUARTER):%H:%M}" for ts in t.index])
        t.insert(1, "Date", [f"{ts:%a %d %b}" for ts in t.index])
        st.dataframe(t.reset_index(drop=True), width="stretch", hide_index=True, height=320)
        st.download_button("Download CSV (full window)", t.to_csv(index=False).encode(),
                           file_name="gr_bmms_live.csv", mime="text/csv")
    st.caption("Charts 1–3 are actual ENTSO-E data (price view fixed to −100…250 €/MWh — spikes run "
               "off the top; drag to zoom, double-click to reset). Chart 4 is the estimated system "
               "state: positive = short (up), negative = long (down). Energy charts 2–4 are in MWh "
               "per 15-min quarter (= MW quarter-average ÷ 4). Energy data lags ~15–40 min.")


def main():
    st.set_page_config(page_title="GR BMMS live", page_icon="🧭", layout="wide")
    with st.sidebar:
        st.header("🧭 GR BMMS live")
        tz_label = st.radio("Clock", ["CET (ENTSO-E platform)", "Greece (Europe/Athens)"],
                            index=0, key="clock_tz")
        window_h = st.select_slider("Window", options=[6, 12, 24, 48], value=12,
                                    format_func=lambda h: f"{h} h", key="window_h")
    if tz_label.startswith("Greece"):
        tz, tz_short = "Europe/Athens", "Greece time"
    else:
        tz, tz_short = "CET", "CET"
    st.title("GR BMMS — mFRR prices, net energy & system state")
    st.caption("Live: mFRR prices, mFRR net and mFRR+aFRR net activated energy (actual data), then "
               "the estimated system state. Auto-updating every minute.")
    live_view(window_h, tz, tz_short)


main()
